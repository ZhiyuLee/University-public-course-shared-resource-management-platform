
# Begin
# auth:lby
# create date:7.18
# description: 爬取武大教务系统

from django.shortcuts import render, redirect
import requests
from bs4 import BeautifulSoup
import os
from openpyxl import workbook  # 写入Excel表所用
from openpyxl import load_workbook  # 读取Excel表所用
import xlrd
from CoursePart import views as CP_views

#爬取武大教务系统并保存通识课程页面的htmldoc
def spider(Cookies):
    path = os.getcwd()
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Cache-Control': 'max-age=0',
        'Connection':'keep-alive',
        'Cookie': 'userLanguage=zh-CN; iPlanetDirectoryPro=PQJnBbaACtXOCvtswr13hi; Hm_lvt_eaa57ca47dacb4ad4f5a257001a3457c=1594816725,1594882063,1594882171,1594882210; JSESSIONID=ED8BB2264659C26136289CE9323DB98F; Hm_lpvt_eaa57ca47dacb4ad4f5a257001a3457c=1594889549',
        'Host': '218.197.149.9',
        'Referer':'http://218.197.149.9/',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.162 Safari/537.36',
        }
    headers['cookie'] = Cookies
    s = requests.Session()
    i = 1
    page_max = 25
    while i <= page_max:
        root = 'http://218.197.149.9/stu/choose_PubLsn_list.jsp?XiaoQu=0&credit=0&keyword=&pageNum='
        url = root + str(i)
        res2 = s.get(url,headers = headers)    
        filename = "spider\\txt\\"+str(i)+'.txt'
        txt_path = os.path.join(path,filename)
        with open(txt_path,'w',encoding='utf-8') as f:
            f.write(res2.text)
            f.close
        i = i + 1

#删除txt中的空格
def delete_space():
    path = os.getcwd()

    i = 1
    page_max = 25
    while i<=page_max:
        filename = "spider\\txt\\"+str(i)+'.txt'
        txt_path = os.path.join(path,filename)
        filename_re = 'spider\\txt\\re'+str(i)+'.txt'
        txt_path_re = os.path.join(path,filename_re)
        with open (txt_path,'r',encoding='utf-8') as f:    
            with open(txt_path_re,'w',encoding='utf-8') as f_re:                    
                for line in f.readlines():
                    data = line.strip()
                    if len(data) != 0 :
                        f_re.write(data)
                        f_re.write('\n')
                f_re.close
            f.close()
        i = i +1

#解析html_doc，获取需要的部分, 并存放进list
def parser(CourseInfo):
    delete_space()
    path = os.getcwd()
    #课程id 课程名	学分 教师名	职称 授课学院 领域 学年 学期 上课时间地点 备注
    CourseID = []
    CourseName = []
    CoursePoint = []
    CourseTeacher = []
    TeacherTitle = []
    Institute = []
    CourseType = []
    CourseYear = []
    CourseTerm = []
    WhereAndWhen = []
    CourseTip = []
    i = 1
    page_max = 25 
    while i<=page_max:   
        filename_re = "spider\\txt\\re"+str(i)+'.txt'
        txt_path_re = os.path.join(path,filename_re)
        with open (txt_path_re,'r',encoding='utf-8') as f:
            html_doc = f.read()
            f.close
        soup = BeautifulSoup(html_doc,'html.parser')
        table_tag = soup.table
        for j in range(3,17):
            if 2*j-1 >= len(table_tag.contents):
                break
            else:
                tr_tag = table_tag.contents[2*j-1]  #基数子节点才是文档，偶数子节点是换行符\n，根据网页分析第一份课程从5开始到33，一页最多15门课程
                CourseID.append(tr_tag.contents[25].input['lessonheadid'])
                CourseName.append(tr_tag.contents[1].text.strip())
                CoursePoint.append(tr_tag.contents[3].text.strip())
                CourseTeacher.append(tr_tag.contents[7].text.strip())
                TeacherTitle.append(tr_tag.contents[9].text.strip())
                Institute.append(tr_tag.contents[11].text.strip())
                CourseType.append(tr_tag.contents[13].text.strip())
                CourseYear.append(tr_tag.contents[15].text.strip())
                CourseTerm.append(tr_tag.contents[17].text.strip())
                WhereAndWhen.append(tr_tag.contents[19].text.strip())      
                CourseTip.append(tr_tag.contents[21].text.strip()) 
                '''
                print(tr_tag.contents[1].text.strip())
                print(tr_tag.contents[3].text.strip())
                print(tr_tag.contents[7].text.strip())
                print(tr_tag.contents[9].text.strip())
                print(tr_tag.contents[11].text.strip())
                print(tr_tag.contents[13].text.strip())
                print(tr_tag.contents[15].text.strip())
                print(tr_tag.contents[17].text.strip())
                print(tr_tag.contents[19].text.strip())
                print(tr_tag.contents[21].text.strip())
                print(tr_tag.contents[25].input['lessonheadid'])
                '''
                
        i = i + 1
    for k in range(len(CourseName)):
        CourseInfo.append([CourseID[k],CourseName[k],CoursePoint[k],CourseTeacher[k],TeacherTitle[k],Institute[k],CourseType[k],CourseYear[k],CourseTerm[k],WhereAndWhen[k],CourseTip[k]])
    return CourseInfo
    #print(CourseInfo)

#导出excel表格
def output_excel():

    wb = workbook.Workbook() #创建Excel对象
    CourseInfoExcel = wb.active #获取当前正在操作的表对象
    #往表中写入标题行，以列表形式
    CourseInfoExcel.append(['课程号','课程名','学分','教师名','职称','授课学院','领域','学年','学期','学时安排','备注'])
    CourseInfo = []
    parser(CourseInfo)
    for i in range(len(CourseInfo)):
        CourseInfoExcel.append(CourseInfo[i])
    wb.save('CourseInfoExcel.xlsx')

def index(request):
    if not request.user.is_superuser:
        return redirect('/admin/')
    if request.method == 'POST':
        Cookies = request.POST.get('Cookies')
        if Cookies!="":
            spider(Cookies)
            output_excel()
        if request.GET.get('update_id')!="":
            import_courses()
    return render(request,'spider.html')

#导入数据库
def import_courses():
    data = xlrd.open_workbook('CourseInfoExcel.xlsx',encoding_override = 'utf-8')
    sheetnames = data.sheet_names()
    sheet_name = sheetnames[0]
    table = data.sheet_by_name(sheet_name)
    nrows = table.nrows
    for i in range(1,nrows - 1):
        if table.row_values(i):
            # 课程名, 课程类别, 授课教师, 授课学院, 课程学分, 学时安排
            CP_views.add_course(table.row_values(i)[1],table.row_values(i)[6],table.row_values(i)[3],table.row_values(i)[4],table.row_values(i)[2],table.row_values(i)[9])
